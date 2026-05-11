#!/usr/bin/env python3
"""
Steam Chart Monitor
SteamSpy API로 상위 1000개 게임을 매일 수집합니다.

수집 데이터:
  - 순위 / 동시접속자(CCU) / 전일대비 CCU 증감
  - 개발사 / 퍼블리셔
  - 장르 / 출시일 / 판매량(추정)
  - 가격 / 할인율
  - 리뷰 수 / 긍정 리뷰 비율

롱런 기준:
  - 1주(7일) / 2주(14일) / 4주(28일) 이상 상위 1000위 유지
"""

import argparse
import re
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import os
import json

# ── 설정 ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH   = os.path.join(SCRIPT_DIR, "steam_chart_monitor.xlsx")
JSON_PATH    = os.path.join(SCRIPT_DIR, "docs", "data.json")
TOP_N        = 1000
LONGRUN_1W   = 7
LONGRUN_2W   = 14
LONGRUN_1M   = 30

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# Steam 연령 제한 우회 쿠키 (18+ / 성인 콘텐츠 게임 appdetails 조회용)
# birthtime=470700000 → 1984년생으로 처리 (42세)
STEAM_COOKIES = {
    "birthtime":            "470700000",
    "lastagecheckage":      "1-0-1984",
    "mature_content":       "1",
    "wants_mature_content": "1",
}


# ── SteamSpy owners 파싱 ──────────────────────────────────────────────────────

def parse_owners_midpoint(owners_str: str) -> int:
    """SteamSpy owners 범위 문자열 → 중간값 (예: '200,000 .. 500,000' → 350000)"""
    try:
        parts = owners_str.replace(",", "").split("..")
        lo = int(parts[0].strip())
        hi = int(parts[1].strip())
        return (lo + hi) // 2
    except Exception:
        return 0


# ── SteamSpy API (상위 게임 목록) ────────────────────────────────────────────

def fetch_steamspy_top(n: int = 1000) -> list:
    """
    SteamSpy all 엔드포인트로 상위 게임 목록 수집.
    page=0 이 최근 2주 플레이어 기준 상위 ~1000개를 반환함.
    결과를 CCU 기준으로 정렬해 순위를 매김.
    """
    games = []
    pages_needed = max(1, (n + 999) // 1000)

    for page in range(pages_needed):
        url = f"https://steamspy.com/api.php?request=all&page={page}"
        print(f"  SteamSpy all 페이지 {page} 수집 중...")
        for attempt in range(3):
            try:
                r = requests.get(url, timeout=30)
                r.raise_for_status()
                data = r.json()

                for appid_str, info in data.items():
                    try:
                        appid = int(appid_str)
                    except ValueError:
                        continue
                    games.append({
                        "appid":      appid,
                        "name_sp":    info.get("name", ""),
                        "ccu":        info.get("ccu", 0) or 0,
                        "developer":  info.get("developer", "") or None,
                        "publisher":  info.get("publisher", "") or None,
                        "genre_sp":   info.get("genre", "") or None,
                        "owners":     info.get("owners", "") or "",
                    })
                print(f"    ✓ {len(data)}개 수집 (누계: {len(games)}개)")
                break
            except Exception as e:
                print(f"    ⚠ 오류 (시도 {attempt+1}/3): {e}")
                time.sleep(5)

        time.sleep(2)

    games.sort(key=lambda x: x["ccu"], reverse=True)
    for i, g in enumerate(games, 1):
        g["rank"] = i

    return games[:n]


# ── Steam Store API ───────────────────────────────────────────────────────────

def fetch_store_details(appid: int) -> dict:
    """Steam Store API: 장르, 출시일, 가격

    filters=basic 은 name 외에 아무것도 반환하지 않으므로 사용하지 않음.
    release_date / genres / price_overview 를 개별 필터로 요청.
    name / developer / publisher 는 SteamSpy 값을 1차 소스로 사용.
    """
    url = (
        f"https://store.steampowered.com/api/appdetails/"
        f"?appids={appid}&cc=kr&filters=release_date,genres,price_overview"
    )
    try:
        r   = requests.get(url, headers=HEADERS, cookies=STEAM_COOKIES, timeout=12)
        raw = r.json().get(str(appid))
        app = raw or {}
        if app.get("success") and app.get("data"):
            d = app["data"]
            p      = d.get("price_overview", {})
            genres = ", ".join(g["description"] for g in d.get("genres", []))
            rd     = d.get("release_date", {})
            # coming_soon=True 이거나 날짜가 없으면 None
            release_date = rd.get("date") if (rd and not rd.get("coming_soon") and rd.get("date")) else None
            return {
                "genres":             genres or None,
                "release_date":       release_date,
                "price_krw":          p.get("final", 0) // 100 if p else None,
                "discount_pct":       p.get("discount_percent", 0) if p else 0,
                "original_price_krw": p.get("initial", 0) // 100 if p else None,
            }
    except Exception:
        pass
    return {
        "genres": None, "release_date": None,
        "price_krw": None, "discount_pct": 0, "original_price_krw": None,
    }


# ── Steam 상점 페이지 개발사/배급사 스크래핑 (SteamSpy 폴백) ──────────────────

def fetch_dev_pub_from_store(appid: int) -> tuple:
    """
    Steam 상점 페이지 HTML에서 개발사/배급사를 직접 파싱.
    SteamSpy에서 값이 없을 때만 호출.
    """
    url = f"https://store.steampowered.com/app/{appid}/"
    try:
        r    = requests.get(url, headers=HEADERS, timeout=15)
        html = r.text

        # 개발자: id="developers_list" 내부 a 태그 텍스트
        developer = None
        dev_match = re.search(
            r'id=["\']developers_list["\'][^>]*>(.*?)</div>', html, re.DOTALL
        )
        if dev_match:
            devs = [d.strip() for d in re.findall(r'>([^<\n]+)</a>', dev_match.group(1)) if d.strip()]
            developer = ", ".join(devs) or None

        # 배급사: "배급사" 또는 "Publisher" 레이블 뒤 div
        publisher = None
        pub_match = re.search(
            r'(?:배급사|Publisher)[^<]*</div>\s*<div[^>]*>(.*?)</div>',
            html, re.DOTALL
        )
        if pub_match:
            pubs = [p.strip() for p in re.findall(r'>([^<\n]+)</a>', pub_match.group(1)) if p.strip()]
            publisher = pubs[0] if pubs else None

        return developer, publisher
    except Exception:
        return None, None


# ── Steam Reviews API ─────────────────────────────────────────────────────────

def fetch_reviews(appid: int) -> dict:
    """Steam Reviews API: 긍정/부정 리뷰 수"""
    url = (
        f"https://store.steampowered.com/appreviews/{appid}"
        f"?json=1&language=all&purchase_type=all&num_per_page=0"
    )
    try:
        r = requests.get(url, headers=HEADERS, timeout=12)
        qs = r.json().get("query_summary", {})
        pos   = qs.get("total_positive", 0) or 0
        neg   = qs.get("total_negative", 0) or 0
        total = pos + neg
        pct   = round(pos / total * 100, 1) if total > 0 else 0
        return {
            "review_score_pct": pct,
            "total_reviews":    total,
        }
    except Exception:
        pass
    return {"review_score_pct": 0, "total_reviews": 0}


# ── 출시 예정 게임 수집 ────────────────────────────────────────────────────────

def _parse_date(date_str: str):
    """날짜 문자열을 date 객체로 파싱.
    지원 형식: '10 May, 2026' / 'May 10, 2026' / '10 May 2026' / '2026-05-10'
    파싱 실패 시 None 반환.
    """
    from datetime import datetime as _dt
    if not date_str:
        return None
    for fmt in ("%d %b, %Y", "%b %d, %Y", "%d %B, %Y", "%B %d, %Y",
                "%d %b %Y", "%B %d %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return _dt.strptime(date_str.strip(), fmt).date()
        except ValueError:
            pass
    return None


def _fetch_steamspy_game_details(appid: int) -> dict:
    """SteamSpy 개별 게임 API → developer/publisher/genre 조회.
    Steam appdetails 실패 시 폴백으로 사용.
    """
    try:
        url = f"https://steamspy.com/api.php?request=appdetails&appid={appid}"
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code == 200:
            d = r.json()
            return {
                "developer": d.get("developer") or None,
                "publisher": d.get("publisher") or None,
                "genres":    d.get("genre") or None,
            }
    except Exception:
        pass
    return {}


# ── Steam 공식 실시간 CCU API ─────────────────────────────────────────────────

def _fetch_one_steam_ccu(appid: int) -> tuple:
    """Steam ISteamUserStats/GetNumberOfCurrentPlayers API로 단일 게임 현재 CCU 조회.
    공식 Steam API이며 인증 불필요. 조회 실패 시 -1 반환.
    """
    url = (
        f"https://api.steampowered.com/ISteamUserStats/"
        f"GetNumberOfCurrentPlayers/v1/?appid={appid}"
    )
    try:
        r = requests.get(url, headers=HEADERS, timeout=8)
        if r.status_code == 200:
            resp = r.json().get("response", {})
            if resp.get("result") == 1:
                return appid, int(resp.get("player_count", 0))
    except Exception:
        pass
    return appid, -1


def fetch_steam_ccu_bulk(appids: list, workers: int = 20) -> dict:
    """Steam 공식 API로 여러 게임의 현재 CCU를 병렬 조회.
    반환: {appid: player_count}  — 조회 실패 appid는 포함되지 않음.
    """
    result = {}
    total  = len(appids)
    print(f"  Steam 공식 CCU 병렬 조회 중... ({total}개, {workers} workers)")
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(_fetch_one_steam_ccu, aid): aid for aid in appids}
        done = 0
        for future in as_completed(futures):
            aid, count = future.result()
            done += 1
            if count >= 0:
                result[aid] = count
            if done % 200 == 0:
                print(f"    [{done}/{total}] CCU 조회 진행 중...")
    ok   = len(result)
    fail = total - ok
    print(f"  ✓ Steam CCU 조회 완료: 성공 {ok}개 / 실패(폴백) {fail}개")
    return result


def fetch_gamalytic_followers(appid: int) -> int:
    """Gamalytic API로 Steam 팔로워 수 조회 (무료, 인증 불필요).
    정상 조회 시 팔로워 수(>=0), 조회 실패 시 -1 반환.
    -1 반환된 게임은 팔로워 필터에서 제외하지 않음.
    """
    try:
        url = f"https://gamalytic.com/api/game-details/{appid}"
        r = requests.get(url, headers=HEADERS, timeout=8)
        if r.status_code == 200:
            val = r.json().get("followers")
            if val is not None:
                return int(val)
    except Exception:
        pass
    return -1


def _enrich_upcoming_item(appid: int, name_fallback: str, header_image_fallback: str = "") -> dict:
    """단일 신작 게임의 상세정보 수집.

    필터 없이 전체 appdetails 응답을 받아 developers/publishers/genres/
    release_date/price_overview 를 모두 가져옴.
    API 실패 시에도 featuredcategories에서 가져온 기본 정보는 반드시 반환.
    """
    cdn_img = f"https://cdn.cloudflare.steamstatic.com/steam/apps/{appid}/header.jpg"
    base = {
        "appid":        appid,
        "name":         name_fallback,
        "developer":    None,
        "publisher":    None,
        "genres":       None,
        "release_date": "",
        "coming_soon":  True,
        "price_krw":    None,
        "discount_pct": 0,
        "header_image": header_image_fallback or cdn_img,
        "followers":    -1,
    }
    # 필터 없이 요청 → developers, publishers, genres, release_date, price_overview 모두 포함
    detail_url = (
        f"https://store.steampowered.com/api/appdetails/"
        f"?appids={appid}&cc=kr"
    )
    # Steam appdetails: 최대 3회 재시도 + 연령 제한 우회 쿠키 포함
    app_data = None
    for attempt in range(3):
        try:
            dr   = requests.get(detail_url, headers=HEADERS,
                                cookies=STEAM_COOKIES, timeout=15)
            raw  = (dr.json() or {}).get(str(appid))  # json() 자체 null 방어
            resp = raw or {}   # 키 값이 null인 경우 방어
            if resp.get("success") and resp.get("data"):
                app_data = resp["data"]
                break
            else:
                print(f"    ⚠ AppID {appid} success=false (시도 {attempt+1}/3)")
                time.sleep(2 ** attempt)  # 1s, 2s, 4s 지수 백오프
        except Exception as e:
            print(f"    ⚠ AppID {appid} 오류 (시도 {attempt+1}/3): {e}")
            time.sleep(2 ** attempt)

    if app_data is None:
        # Steam API 완전 실패 → SteamSpy 폴백으로 developer/publisher/genre 보완
        print(f"    ⚠ AppID {appid} Steam API 실패, SteamSpy 폴백 시도")
        sp = _fetch_steamspy_game_details(appid)
        if sp:
            base.update({k: v for k, v in sp.items() if v is not None})
            print(f"      SteamSpy 폴백 성공: dev={sp.get('developer')}, genre={sp.get('genres')}")
        return base

    d = app_data

    # 개발사 / 배급사
    developer = ", ".join(d.get("developers", [])) or None
    publisher = ", ".join(d.get("publishers", [])) or None

    # 장르
    genres = ", ".join(g["description"] for g in d.get("genres", [])) or None

    # 출시일 / coming_soon
    rd = d.get("release_date", {}) or {}
    release_date_str = rd.get("date", "") or ""
    coming_soon_flag = bool(rd.get("coming_soon", True))

    # 가격 (₩)
    p = d.get("price_overview") or {}
    is_free   = d.get("is_free", False)
    price_krw = 0 if is_free else (int(p.get("final", 0)) // 100 if p else None)
    disc_pct  = int(p.get("discount_percent", 0)) if p else 0

    # 헤더 이미지: API > featuredcategories > CDN 순
    header_image = d.get("header_image") or header_image_fallback or cdn_img

    base.update({
        "name":         d.get("name") or name_fallback,
        "developer":    developer,
        "publisher":    publisher,
        "genres":       genres,
        "release_date": release_date_str,
        "coming_soon":  coming_soon_flag,
        "price_krw":    price_krw,
        "discount_pct": disc_pct,
        "header_image": header_image,
    })
    return base


def _fetch_upcoming_search(seen: set, date_from, date_to) -> list:
    """Steam 검색 API로 출시 예정 게임 수집.

    filter=comingsoon 파라미터로 Steam 전체 upcoming 게임을 가져옴.
    featuredcategories 의 소수 큐레이션 게임 외 실제 출시 예정작을 보완.
    """
    items = []
    url = (
        "https://store.steampowered.com/search/results/"
        "?sort_by=Released_ASC&filter=comingsoon"
        "&category1=998&count=100&cc=kr&l=koreana&json=1"
    )
    try:
        r    = requests.get(url, headers=HEADERS, timeout=15)
        data = r.json()
        raw  = data.get("items", [])
        print(f"  [Steam 검색 upcoming] {len(raw)}개")
        for it in raw:
            aid = it.get("id") or it.get("appid")
            if not aid or aid in seen:
                continue
            # release_str 예: "11 May, 2026" — 범위 내인지 사전 확인
            rel_str = (it.get("release_date") or "").strip()
            parsed  = _parse_date(rel_str) if rel_str else None
            if parsed is not None and not (date_from <= parsed <= date_to):
                continue   # 범위 밖이면 appdetails 호출 낭비 방지
            seen.add(aid)
            items.append({
                "id":           aid,
                "name":         it.get("name", ""),
                "header_image": it.get("logo", "") or it.get("header_image", ""),
            })
    except Exception as e:
        print(f"  ⚠ Steam 검색 upcoming 실패: {e}")
    return items


def fetch_upcoming_games() -> list:
    """
    Steam 신작 캘린더 수집.

    소스 ①  featuredcategories API — Steam 큐레이션 coming_soon / new_releases
    소스 ②  Steam 검색 API (filter=comingsoon) — 전체 출시 예정작 (최대 100개)

    두 소스를 합산·중복 제거 후 각 게임의 appdetails 수집.
    - Gamalytic API로 팔로워 수 조회 → 500 미만 게임 제외
    - 팔로워 조회 실패(-1) 게임은 포함 (불명 처리)
    - 10,000+ 팔로워 게임은 HTML 대시보드에서 강조 표시
    """
    print("▶ 신작 캘린더 수집 중...")
    today     = date.today()
    date_from = today - timedelta(days=7)   # 최근 1주 출시 포함
    date_to   = today + timedelta(days=60)  # 이번 달 + 다음 달까지 예정작 포함

    seen  = set()
    items = []

    # ── 소스 ①: featuredcategories ─────────────────────────────────────────
    cat_url = "https://store.steampowered.com/api/featuredcategories/?cc=kr&l=koreana"
    try:
        r    = requests.get(cat_url, headers=HEADERS, timeout=15)
        cats = r.json()

        for section in ("coming_soon", "new_releases"):
            sec_items = cats.get(section, {}).get("items", [])
            print(f"  [featuredcategories/{section}] {len(sec_items)}개")
            for it in sec_items:
                aid = it.get("id")
                if aid and aid not in seen:
                    seen.add(aid)
                    items.append({
                        "id":           aid,
                        "name":         it.get("name", ""),
                        "header_image": it.get("header_image", ""),
                    })
    except Exception as e:
        print(f"  ⚠ featuredcategories 수집 실패: {e}")

    # ── 소스 ②: Steam 검색 API (coming soon 전체) ──────────────────────────
    search_items = _fetch_upcoming_search(seen, date_from, date_to)
    items.extend(search_items)

    print(f"  총 {len(items)}개 (중복 제거 후) → 상세 + 팔로워 수집 중...")
    games = []
    for it in items:
        # 1. appdetails 수집 (개발사/퍼블리셔/장르/가격/출시일)
        g = _enrich_upcoming_item(it["id"], it["name"], it.get("header_image", ""))

        # 2. 날짜 범위 필터: 파싱 성공 시에만 적용 (파싱 실패는 포함)
        parsed = _parse_date(g.get("release_date", ""))
        if parsed is not None and not (date_from <= parsed <= date_to):
            print(f"    날짜 범위 외 제외: {g['name']} ({g['release_date']})")
            time.sleep(0.2)
            continue

        # 3. Gamalytic 팔로워 조회
        followers = fetch_gamalytic_followers(it["id"])
        g["followers"] = followers
        print(f"    {g['name']} | 팔로워={followers if followers >= 0 else '조회불가'}")

        # 4. 팔로워 500 미만 제외 (-1=조회실패는 포함)
        if 0 <= followers < 500:
            print(f"    팔로워 부족 제외 ({followers}명)")
            time.sleep(0.2)
            continue

        games.append(g)
        time.sleep(0.5)

    # 날짜 오름차순 정렬 (출시 예정 먼저, 이후 최근 출시)
    games.sort(key=lambda x: (not x["coming_soon"], x.get("release_date") or ""))
    print(f"  ✓ 최종 {len(games)}개 수집 완료 (팔로워 500+ 또는 조회 불명)")
    return games


# ── 전일 대비 CCU 증감 계산 ───────────────────────────────────────────────────

def add_ccu_change(today_df: pd.DataFrame, existing_df: pd.DataFrame) -> pd.DataFrame:
    """오늘 CCU/순위와 전일 CCU/순위를 비교하여 증감 컬럼 추가.

    추가 컬럼:
      ccu_change      — 전일 대비 CCU 절대 증감 (Steam 공식 실시간 CCU 기준)
      ccu_change_pct  — CCU 변화율(%)
      rank_change     — 전일 대비 순위 변화 (양수 = 상승, 음수 = 하락)
                        예) 전일 7위 → 오늘 4위: rank_change = +3
    """
    today_df = today_df.copy()

    def _set_null_changes(df):
        df["ccu_change"]     = None
        df["ccu_change_pct"] = None
        df["rank_change"]    = None
        return df

    if existing_df.empty:
        print("  ⚠ 기존 데이터 없음 → 전일 증감 계산 불가 (null)")
        return _set_null_changes(today_df)

    today_str  = date.today().isoformat()
    prev_dates = existing_df[existing_df["date"] != today_str]["date"].unique()

    if len(prev_dates) == 0:
        print("  ⚠ 전일 데이터 없음 (Excel에 오늘 데이터만 존재) → null")
        return _set_null_changes(today_df)

    prev_date = max(prev_dates)
    print(f"  ✓ 전일 기준일: {prev_date}  (누적 날짜 수: {len(prev_dates)}일)")

    # appid 타입 통일 (int) — Excel float64 잔재 방어
    today_df["appid"] = today_df["appid"].astype(int)

    prev_ccu_df = (
        existing_df[existing_df["date"] == prev_date][["appid", "ccu"]]
        .copy()
        .rename(columns={"ccu": "ccu_prev"})
    )
    prev_ccu_df["appid"] = prev_ccu_df["appid"].astype(int)

    prev_rank_df = (
        existing_df[existing_df["date"] == prev_date][["appid", "rank"]]
        .copy()
        .rename(columns={"rank": "rank_prev"})
    )
    prev_rank_df["appid"] = prev_rank_df["appid"].astype(int)

    merged = today_df.merge(prev_ccu_df, on="appid", how="left")
    merged = merged.merge(prev_rank_df,  on="appid", how="left")

    ccu_matched  = int(merged["ccu_prev"].notna().sum())
    rank_matched = int(merged["rank_prev"].notna().sum())
    print(f"  ✓ 전일 매칭: CCU {ccu_matched}/{len(merged)}개 / 순위 {rank_matched}/{len(merged)}개")

    # ── CCU 증감 (float64 — Int64/pd.NA 는 openpyxl 오류 유발) ──────────────
    ccu_mask = merged["ccu_prev"].notna()
    merged["ccu_change"] = (merged["ccu"] - merged["ccu_prev"]).where(ccu_mask)
    merged["ccu_change_pct"] = (
        (merged["ccu"] - merged["ccu_prev"]) / merged["ccu_prev"] * 100
    ).where(ccu_mask).round(1)

    # ── 순위 증감 (양수=상승, 음수=하락) ───────────────────────────────────
    rank_mask = merged["rank_prev"].notna()
    merged["rank_change"] = (merged["rank_prev"] - merged["rank"]).where(rank_mask)

    return merged.drop(columns=["ccu_prev", "rank_prev"])


# ── 데이터 수집 ───────────────────────────────────────────────────────────────

def collect_today_data() -> pd.DataFrame:
    print("▶ SteamSpy API 상위 1000 게임 수집 중...")
    sp_games = fetch_steamspy_top(TOP_N)
    print(f"  총 {len(sp_games)}개 게임 수집 완료")

    today_str = date.today().isoformat()
    rows = []

    for i, g in enumerate(sp_games, 1):
        print(f"  [{i:4d}/{len(sp_games)}] AppID {g['appid']} ({g['name_sp'][:30]})")

        store   = fetch_store_details(g["appid"])
        reviews = fetch_reviews(g["appid"])
        time.sleep(0.5)

        name = g["name_sp"]  # Steam Store API는 이제 name 미반환, SteamSpy 이름 사용

        # 개발사/퍼블리셔: SteamSpy 1차 → 없으면 스팀 상점 페이지 스크래핑 폴백
        developer = g.get("developer")
        publisher = g.get("publisher")
        if not developer and not publisher:
            developer, publisher = fetch_dev_pub_from_store(g["appid"])
            time.sleep(0.3)
        # 장르: Steam API를 1차, SteamSpy를 보조로
        genres = store["genres"] or g.get("genre_sp")

        # 판매량 추정: SteamSpy owners 중간값
        owners_estimate = parse_owners_midpoint(g.get("owners", "")) or None

        rows.append({
            "date":               today_str,
            "rank":               g["rank"],
            "appid":              g["appid"],
            "name":               name,
            "developer":          developer,
            "publisher":          publisher,
            "genres":             genres,
            "release_date":       store["release_date"],
            "owners_estimate":    owners_estimate,
            "ccu":                g["ccu"],
            "review_score_pct":   reviews["review_score_pct"],
            "total_reviews":      reviews["total_reviews"],
            "price_krw":          store["price_krw"],
            "discount_pct":       store["discount_pct"],
            "original_price_krw": store["original_price_krw"],
        })

    df = pd.DataFrame(rows)

    # ── Steam 공식 실시간 CCU로 교체 (SteamSpy CCU는 다주간 평균이라 변동 미미) ──
    print("\n▶ Steam 공식 실시간 CCU 조회 중...")
    steam_ccu = fetch_steam_ccu_bulk(df["appid"].tolist())
    df["ccu_steamspy"] = df["ccu"]   # SteamSpy 원본 보존
    df["ccu"] = df["appid"].map(steam_ccu)
    # 조회 실패한 게임은 SteamSpy CCU 폴백
    fallback_mask = df["ccu"].isna()
    df.loc[fallback_mask, "ccu"] = df.loc[fallback_mask, "ccu_steamspy"]
    df["ccu"] = df["ccu"].fillna(0).astype(int)
    n_real     = int((~fallback_mask).sum())
    n_fallback = int(fallback_mask.sum())
    print(f"  ✓ 실시간 CCU 적용: {n_real}개 / SteamSpy 폴백: {n_fallback}개")

    # ── Steam 실시간 CCU 기준으로 순위 재정렬 ──────────────────────────────
    # SteamSpy 순서(ccu_steamspy 기준)를 Steam 실시간 CCU 기준으로 교체
    df = df.sort_values("ccu", ascending=False).reset_index(drop=True)
    df["rank"] = range(1, len(df) + 1)
    print(f"  ✓ Steam 실시간 CCU 기준 순위 재정렬 완료")

    return df


# ── 롱런 분석 ─────────────────────────────────────────────────────────────────

def analyze_longrun(df: pd.DataFrame, min_days: int) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["date"] = pd.to_datetime(df["date"])

    agg_dict = dict(
        days_in_top         = ("date",             "nunique"),
        avg_rank            = ("rank",             "mean"),
        best_rank           = ("rank",             "min"),
        developer           = ("developer",        "last"),
        publisher           = ("publisher",        "last"),
        genres              = ("genres",           "last"),
        release_date        = ("release_date",     "last"),
        avg_ccu             = ("ccu",              "mean"),
        latest_ccu          = ("ccu",              "last"),
        avg_review_score    = ("review_score_pct", "mean"),
        latest_review_score = ("review_score_pct", "last"),
        total_reviews       = ("total_reviews",    "last"),
        latest_price        = ("price_krw",        "last"),
        max_discount        = ("discount_pct",     "max"),
        first_seen          = ("date",             "min"),
        last_seen           = ("date",             "max"),
    )
    if "owners_estimate" in df.columns:
        agg_dict["owners_estimate"] = ("owners_estimate", "last")

    stats = df.groupby(["appid", "name"]).agg(**agg_dict).reset_index()

    result = stats[stats["days_in_top"] >= min_days].copy()
    result.sort_values("days_in_top", ascending=False, inplace=True)

    result["avg_rank"]         = result["avg_rank"].round(1)
    result["avg_ccu"]          = result["avg_ccu"].round(0).astype(int)
    result["avg_review_score"] = result["avg_review_score"].round(1)
    result["first_seen"]       = result["first_seen"].dt.strftime("%Y-%m-%d")
    result["last_seen"]        = result["last_seen"].dt.strftime("%Y-%m-%d")

    return result


# ── Excel 출력 ────────────────────────────────────────────────────────────────

def _xval(v):
    """pandas pd.NA / np.nan → None 변환 (openpyxl 호환)
    Int64 nullable integer 컬럼의 <NA> 값이 그대로 셀에 들어가면 ValueError 발생.
    """
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL  = PatternFill("solid", start_color="EBF3FB")
DIS_FILL  = PatternFill("solid", start_color="C6EFCE")
LRN1_FILL = PatternFill("solid", start_color="E8F5E9")
LRN2_FILL = PatternFill("solid", start_color="FFF3CD")
LRN4_FILL = PatternFill("solid", start_color="FFD700")
UP_FILL   = PatternFill("solid", start_color="E8F5E9")
UPC_FILL  = PatternFill("solid", start_color="F3E5F5")
CENTER    = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(all_df, today_df, lr1, lr2, lr1m, upcoming):
    wb = Workbook()

    # ── 시트 1: 일별 스냅샷 ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "일별 스냅샷"
    SNAP_COLS = {
        "date":               "날짜",
        "rank":               "순위",
        "rank_change":        "순위변동",
        "appid":              "AppID",
        "name":               "게임명",
        "developer":          "개발사",
        "publisher":          "퍼블리셔",
        "genres":             "장르",
        "release_date":       "출시일",
        "owners_estimate":    "판매량(추정)",
        "ccu":                "동접자(실시간)",
        "ccu_steamspy":       "동접자(SteamSpy)",
        "ccu_change":         "전일증감",
        "ccu_change_pct":     "증감(%)",
        "review_score_pct":   "긍정리뷰(%)",
        "total_reviews":      "리뷰수",
        "price_krw":          "가격(₩)",
        "discount_pct":       "할인(%)",
        "original_price_krw": "정가(₩)",
    }
    ws1.append(list(SNAP_COLS.values()))
    _style_header(ws1, 1, len(SNAP_COLS))
    keys = list(SNAP_COLS.keys())
    for ri, row in enumerate(all_df.itertuples(index=False), 2):
        for ci, k in enumerate(keys, 1):
            ws1.cell(ri, ci, value=_xval(getattr(row, k, None)))
        if ri % 2 == 0:
            for ci in range(1, len(keys) + 1):
                ws1.cell(ri, ci).fill = ALT_FILL
    _set_col_widths(ws1, [12, 5, 8, 10, 34, 24, 24, 22, 12, 14, 14, 14, 10, 8, 10, 10, 10, 8, 10])
    ws1.freeze_panes = "A2"

    # ── 시트 2: 오늘의 차트 ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("오늘의 차트")
    ws2["A1"] = f"Steam 인기 차트 — {date.today().isoformat()}"
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws2.append([])
    T_COLS = ["순위", "게임명", "개발사", "퍼블리셔", "장르", "출시일", "판매량(추정)",
              "동접자", "전일증감(%)", "리뷰수(긍정%)", "가격(₩)", "할인율(%)"]
    T_KEYS = ["rank", "name", "developer", "publisher", "genres", "release_date",
              "owners_estimate", "ccu", "_ccu_change", "_review_display",
              "price_krw", "discount_pct"]
    ws2.append(T_COLS)
    _style_header(ws2, 3, len(T_COLS))
    for ri, row in enumerate(today_df.itertuples(index=False), 4):
        for ci, k in enumerate(T_KEYS, 1):
            if k == "_ccu_change":
                chg = getattr(row, "ccu_change", None)
                pct = getattr(row, "ccu_change_pct", None)
                if chg is not None and pct is not None:
                    sign = "▲" if chg > 0 else ("▼" if chg < 0 else "")
                    val  = f"{sign} {chg:+,} ({pct:+.1f}%)" if chg != 0 else "±0"
                else:
                    val = ""
                ws2.cell(ri, ci, value=val)
            elif k == "_review_display":
                pct = _xval(getattr(row, "review_score_pct", 0)) or 0
                cnt = _xval(getattr(row, "total_reviews", 0)) or 0
                ws2.cell(ri, ci, value=f"{cnt:,} ({pct}%)" if cnt else "")
            else:
                ws2.cell(ri, ci, value=_xval(getattr(row, k, None)))
        disc = getattr(row, "discount_pct", 0) or 0
        chg  = getattr(row, "ccu_change", None)
        if disc > 0:
            fill = DIS_FILL
        elif chg is not None and chg > 0:
            fill = UP_FILL
        elif ri % 2 == 0:
            fill = ALT_FILL
        else:
            fill = None
        if fill:
            for ci in range(1, len(T_KEYS) + 1):
                ws2.cell(ri, ci).fill = fill
    _set_col_widths(ws2, [5, 34, 24, 24, 22, 12, 14, 12, 16, 18, 10, 8])
    ws2.freeze_panes = "A4"

    # ── 시트 3~5: 롱런 분석 ─────────────────────────────────────────────────
    LR_COLS = {
        "name":                "게임명",
        "days_in_top":         "유지 일수",
        "avg_rank":            "평균 순위",
        "best_rank":           "최고 순위",
        "developer":           "개발사",
        "publisher":           "퍼블리셔",
        "genres":              "장르",
        "release_date":        "출시일",
        "owners_estimate":     "판매량(추정)",
        "avg_ccu":             "평균 동접",
        "latest_ccu":          "최근 동접",
        "avg_review_score":    "평균 긍정리뷰(%)",
        "latest_review_score": "최근 긍정리뷰(%)",
        "total_reviews":       "총 리뷰수",
        "latest_price":        "현재 가격(₩)",
        "max_discount":        "최대 할인(%)",
        "first_seen":          "첫 관측일",
        "last_seen":           "최근 관측일",
    }
    LR_WIDTHS = [34, 10, 10, 10, 24, 24, 22, 12, 14, 12, 12, 16, 16, 12, 14, 10, 14, 14]

    def write_lr(ws, title, df, fill, min_days):
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=13, color="B8860B")
        ws.append([])
        ws.append(list(LR_COLS.values()))
        _style_header(ws, 3, len(LR_COLS))
        if not df.empty:
            for ri, row in enumerate(df.itertuples(index=False), 4):
                for ci, k in enumerate(LR_COLS.keys(), 1):
                    ws.cell(ri, ci, value=_xval(getattr(row, k, None)))
                for ci in range(1, len(LR_COLS) + 1):
                    ws.cell(ri, ci).fill = fill
        else:
            ws["A4"] = f"아직 {min_days}일 분량의 데이터가 쌓이지 않았습니다."
            ws["A4"].font = Font(italic=True, color="888888")
        _set_col_widths(ws, LR_WIDTHS)
        ws.freeze_panes = "A4"

    write_lr(wb.create_sheet("스테디셀러 7일+"),
             f"스테디셀러 — 7일+ 상위 1000위 유지 — {date.today().isoformat()}", lr1, LRN1_FILL, LONGRUN_1W)
    write_lr(wb.create_sheet("스테디셀러 14일+"),
             f"스테디셀러 — 14일+ 상위 1000위 유지 — {date.today().isoformat()}", lr2, LRN1_FILL, LONGRUN_2W)
    write_lr(wb.create_sheet("장기흥행 30일+"),
             f"장기 흥행 — 30일+ 상위 1000위 유지 — {date.today().isoformat()}", lr1m, LRN4_FILL, LONGRUN_1M)

    # ── 시트 6: 신작 캘린더 ─────────────────────────────────────────────────
    ws6 = wb.create_sheet("신작 캘린더")
    ws6["A1"] = f"Steam 주목 출시 예정 게임 — {date.today().isoformat()}"
    ws6["A1"].font = Font(bold=True, size=13, color="6A1B9A")
    ws6.append([])
    UPC_COLS = ["게임명", "개발사", "퍼블리셔", "장르", "팔로워", "출시예정일", "상태", "가격(₩)", "AppID"]
    ws6.append(UPC_COLS)
    _style_header(ws6, 3, len(UPC_COLS))
    if upcoming:
        for ri, g in enumerate(upcoming, 4):
            fw = g.get("followers", -1)
            vals = [
                g.get("name"),
                g.get("developer"),
                g.get("publisher"),
                g.get("genres"),
                fw if fw >= 0 else None,
                g.get("release_date"),
                "출시 예정" if g.get("coming_soon") else "출시됨",
                g.get("price_krw"),
                g.get("appid"),
            ]
            for ci, v in enumerate(vals, 1):
                ws6.cell(ri, ci, value=v)
            for ci in range(1, len(UPC_COLS) + 1):
                ws6.cell(ri, ci).fill = UPC_FILL
    else:
        ws6["A4"] = "출시 예정 게임 데이터를 가져오지 못했습니다."
        ws6["A4"].font = Font(italic=True, color="888888")
    _set_col_widths(ws6, [34, 24, 24, 22, 10, 15, 10, 10, 10])
    ws6.freeze_panes = "A4"

    wb.save(EXCEL_PATH)
    print(f"✔ Excel 저장: {EXCEL_PATH}")


# ── JSON 출력 ─────────────────────────────────────────────────────────────────

def write_json(today_df, lr1, lr2, lr1m, upcoming, accumulated_days: int = 0):
    def to_records(df, cols=None):
        if df.empty:
            return []
        d = df[cols].copy() if cols else df.copy()
        # float NaN / Int64 pd.NA → null (pandas to_json이 자동 처리)
        return json.loads(d.to_json(orient="records", force_ascii=False))

    TODAY_COLS = [
        "rank", "rank_change", "appid", "name", "developer", "publisher",
        "genres", "release_date", "owners_estimate",
        "ccu", "ccu_change", "ccu_change_pct",
        "review_score_pct", "total_reviews",
        "price_krw", "discount_pct",
    ]
    # ccu_change 유효 비율 계산 (진단용)
    ccu_valid = int(today_df["ccu_change"].notna().sum()) if "ccu_change" in today_df.columns else 0
    ccu_total = len(today_df)

    data = {
        "updated":          date.today().isoformat(),
        "accumulated_days": accumulated_days,       # 누적 수집 일수
        "ccu_change_valid": ccu_valid,              # 전일 증감 계산된 게임 수
        "ccu_change_total": ccu_total,              # 전체 게임 수
        "today_chart":      to_records(today_df, TODAY_COLS),
        "longrun_1w":       to_records(lr1),
        "longrun_2w":       to_records(lr2),
        "longrun_1m":       to_records(lr1m),
        "upcoming_games":   upcoming,
    }
    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    print(f"✔ JSON 저장: {JSON_PATH} (upcoming: {len(upcoming)}개)")


# ── 기존 데이터 로드 ──────────────────────────────────────────────────────────

def load_existing() -> pd.DataFrame:
    if not os.path.exists(EXCEL_PATH):
        return pd.DataFrame()
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="일별 스냅샷")
        # 일별 스냅샷 헤더는 한글(날짜, 동접자...)로 저장되어 있음.
        # 영문 컬럼명으로 역매핑해야 add_ccu_change / analyze_longrun이 작동함.
        KR_TO_EN = {
            "날짜": "date", "순위": "rank", "순위변동": "rank_change", "AppID": "appid",
            "게임명": "name", "개발사": "developer", "퍼블리셔": "publisher",
            "장르": "genres", "출시일": "release_date",
            "판매량(추정)": "owners_estimate",
            # 컬럼명 변경 전후 모두 지원 (하위 호환)
            "동접자": "ccu", "동접자(실시간)": "ccu",
            "동접자(SteamSpy)": "ccu_steamspy",
            "전일증감": "ccu_change", "증감(%)": "ccu_change_pct",
            "긍정리뷰(%)": "review_score_pct", "리뷰수": "total_reviews",
            "가격(₩)": "price_krw", "할인(%)": "discount_pct",
            "정가(₩)": "original_price_krw",
        }
        df = df.rename(columns=KR_TO_EN)
        # Excel에서 읽은 appid는 float64로 저장됨 → int로 변환해야 merge가 정상 작동
        if "appid" in df.columns:
            df["appid"] = pd.to_numeric(df["appid"], errors="coerce").fillna(0).astype(int)
        df["date"] = pd.to_datetime(df["date"]).dt.date.astype(str)
        print(f"  ✓ 기존 데이터 로드: {len(df)}행 ({df['date'].nunique()}일치)")
        return df
    except Exception as e:
        print(f"기존 파일 로드 실패 ({e}), 새로 시작합니다.")
        return pd.DataFrame()


# ── 메인 ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Steam Chart Monitor")
    parser.add_argument(
        "--dry-run", action="store_true",
        help="데이터를 수집하지만 Excel/JSON에 저장하지 않음 (테스트용)"
    )
    args = parser.parse_args()
    dry_run = args.dry_run

    today_str = date.today().isoformat()
    print(f"\n{'='*60}")
    print(f"  Steam Chart Monitor  |  {today_str}")
    if dry_run:
        print(f"  ⚠ DRY-RUN 모드 — 수집만 하고 저장하지 않습니다")
    print(f"{'='*60}")

    # 1. 오늘 데이터 수집
    today_df = collect_today_data()

    # 2. 기존 데이터 로드
    existing = load_existing()
    if not existing.empty:
        existing = existing[existing["date"] != today_str]

    # 3. 전일대비 CCU 증감 계산
    today_df = add_ccu_change(today_df, existing)

    # 4. 전체 데이터 합치기
    all_df = (
        pd.concat([existing, today_df], ignore_index=True)
        if not existing.empty
        else today_df
    )

    # 5. 롱런 분석
    lr1  = analyze_longrun(all_df.copy(), LONGRUN_1W)
    lr2  = analyze_longrun(all_df.copy(), LONGRUN_2W)
    lr1m = analyze_longrun(all_df.copy(), LONGRUN_1M)
    print(f"\n▶ 스테디셀러 (7일+): {len(lr1)}개")
    print(f"▶ 스테디셀러 (14일+): {len(lr2)}개")
    print(f"▶ 장기 흥행 (30일+): {len(lr1m)}개")

    # 6. 출시 예정 게임 수집
    upcoming = fetch_upcoming_games()

    # 7. 저장
    acc_days = int(all_df["date"].nunique()) if not all_df.empty else 1
    print(f"\n▶ 누적 데이터: {acc_days}일치 ({len(all_df)}행)")

    if dry_run:
        print("\n⚠ DRY-RUN 모드 — Excel/JSON 저장을 건너뜁니다.")
        print("  (실제 저장하려면 --dry-run 없이 실행하세요)\n")
    else:
        build_excel(all_df, today_df, lr1, lr2, lr1m, upcoming)
        write_json(today_df, lr1, lr2, lr1m, upcoming, accumulated_days=acc_days)
        print("  완료!\n")


if __name__ == "__main__":
    main()
